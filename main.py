from fastapi import FastAPI, HTTPException, Depends, Request, APIRouter, Query
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from rcc_app import models, schemas, crud
from collections import defaultdict
from .database import engine, Base, get_db, get_db_access, get_db_prod_stats
from datetime import datetime,timedelta,time 
import pyodbc
import socket
import sys
import os

from openpyxl import load_workbook
from typing import List, Generator
from collections import defaultdict


sys.path.append(os.path.dirname(os.path.abspath(__file__)))

Base.metadata.create_all(bind=engine)

app = FastAPI()


# Enable CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# LOGIN
@app.post("/login/")
def login(login_data: schemas.Login, request: Request, db: Session = Depends(get_db)):
    # Get the IP address of the client
    client_ip = request.client.host

    # Query user by email
    user = db.query(models.User).filter(models.User.usremail == login_data.email).first()
    # Handle case where the user does not exist
    if not user:
        # Log the failed login attempt
        log_attempt = models.LoginAttempt(
            usrid=None,  # Set to None when the user does not exist
            ipaddr=client_ip,
            attemptat=datetime.now(),
            success=False,
            reason="User not found"
        )
        db.add(log_attempt)
        db.commit()
        raise HTTPException(status_code=401, detail="Invalid email or password")

    # Validate the password
    if user.__getattribute__("password") != login_data.password:  # Use __getattribute__ to access the field
        # Log the failed login attempt
        log_attempt = models.LoginAttempt(
            usrid=user.usrid,
            ipaddr=client_ip,
            attemptat=datetime.now(),
            success=False,
            reason="Invalid password"
        )
        db.add(log_attempt)
        db.commit()
        raise HTTPException(status_code=401, detail="Invalid email or password")

    # Log the successful login attempt
    log_attempt = models.LoginAttempt(
        usrid=user.usrid,
        ipaddr=client_ip,
        attemptat=datetime.now(),
        success=True,
        reason="Successful Login"
    )
    db.add(log_attempt)
    db.commit()

    return {"message": "Login successful", "user": {"id": user.usrid, "name": user.usrnamedisplay}}

# fetching users from postgre
@app.get("/user/{user_id}")
def get_user_info(user_id: int, db: Session = Depends(get_db)):
    """
    User information retrieval endpoint
    """
    user = db.query(models.User).filter(models.User.usrid == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return {"id": user.usrid, "name": user.usrnamedisplay}

#REPORTS
#RCC Reports --------------------
#Summary Report 
# offline wtgs heading
@app.get("/stoppage_headings")
def get_stoppage_legend(
    startdate: str = Query(..., description="Start date in YYYY-MM-DD"),
    enddate: str = Query(..., description="End date in YYYY-MM-DD"),
    db: pyodbc.Connection = Depends(get_db_access)
):
    cursor = db.cursor()

    try:
        start_dt = datetime.strptime(startdate, "%Y-%m-%d")
        end_dt = datetime.strptime(enddate, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    cursor.execute(
        """
        SELECT 
            COUNT(IIf(r.rtnName NOT LIKE '*Communication*' 
                AND r.rtnName NOT LIKE '*IDF Fault*' 
                AND r.rtnName NOT LIKE '*IDF Outage*', 1, NULL)) AS total_stoppages,
            COUNT(IIf(r.rtnName IN ('Schedule Service', 'Scheduled - Adhoc', 'Scheduled Inspections'), 1, NULL)) AS scheduled_stoppages,
            COUNT(IIf(r.rtnName NOT IN ('Schedule Service', 'Scheduled - Adhoc', 
                'Scheduled Inspections', 'Fault', 'IDF Fault', 'Communication'), 1, NULL)) AS non_scheduled_stoppages,
            COUNT(IIf(r.rtnName IN ('Fault','IDF Fault'), 1, NULL)) AS fault_stoppages,
            ROUND(AVG(IIf(e.dtTS7DownFinish IS NOT NULL AND e.dtTS3MaintBegin IS NOT NULL, 
                DateDiff('s', e.dtTS3MaintBegin, e.dtTS7DownFinish) / 3600.0, 0)), 2) AS avg_maint,
            ROUND(AVG(IIf(e.dtTS7DownFinish IS NOT NULL AND e.dtTS1DownBegin IS NOT NULL, 
                DateDiff('s', e.dtTS1DownBegin, e.dtTS7DownFinish) / 3600.0, 0)), 2) AS avg_down
        FROM 
            (tblEvent AS e
            INNER JOIN tblRationale AS r ON e.rtnID = r.rtnID)
        WHERE 
            e.dtTS1DownBegin BETWEEN ? AND ?
        """,
        (start_dt, end_dt)
    )

    rows = cursor.fetchall()

    # Assuming the query will return one row with the count result
    if rows:
        total_stoppages = rows[0][0]
        scheduled_stoppages = rows[0][1]
        non_scheduled_stoppages = rows[0][2]
        fault_stoppages = rows[0][3]
        avg_maint = rows[0][4]
        avg_down = rows[0][5]
    else:
        total_stoppages = 0
        scheduled_stoppages = 0
        non_scheduled_stoppages = 0
        fault_stoppages = 0
        avg_maint = 0
        avg_down = 0

    # Structure the result into a dictionary or list format
    result = {
        "total_stoppages": total_stoppages,
        "scheduled_stoppages": scheduled_stoppages,
        "non_scheduled_stoppages": non_scheduled_stoppages,
        "fault_stoppages": fault_stoppages,
        "avg_maint": avg_maint,
        "avg_down": avg_down
    }

    return result

# 2 charts
@app.get("/summary_stoppages")
def get_summary_stoppages(
    startdate: str = Query(..., description="Start date in YYYY-MM-DD"),
    enddate: str = Query(..., description="End date in YYYY-MM-DD"),
    db: pyodbc.Connection = Depends(get_db_access)
):
    cursor = db.cursor()

    try:
        start_dt = datetime.strptime(startdate, "%Y-%m-%d")
        end_dt = datetime.strptime(enddate, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    cursor.execute("""
        SELECT 
            f.facABBR AS windfarm,
            r.rtnName AS category,
            e.dtTS1DownBegin AS stop_time,
            e.dtTS3MaintBegin AS maint_time,
            e.dtTS7EventFinish AS start_time
        FROM 
            ((tblEvent AS e
            INNER JOIN tblFacility AS f ON e.facID = f.facID)
            INNER JOIN tblRationale AS r ON e.rtnID = r.rtnID)
        WHERE       
            e.dtTS1DownBegin BETWEEN ? AND ?
    """, (start_dt, end_dt))

    rows = cursor.fetchall()

    summary = defaultdict(lambda: defaultdict(int))
    downtime_data = defaultdict(list)
    service_data = defaultdict(list)

    for row in rows:
        wf = row.windfarm
        cat = row.category.strip().lower() if row.category else ""
        summary[wf]["Total Stops"] += 1

        if cat == "schedule service":
            summary[wf]["Scheduled Services"] += 1
        elif cat in ["fault", "idf fault"]:
            summary[wf]["Faults"] += 1
        else:
            summary[wf]["Non Scheduled Services"] += 1

        if row.start_time and row.stop_time:
            dt = (row.start_time - row.stop_time).total_seconds() / 3600
            downtime_data[wf].append(dt)

        if row.maint_time and row.start_time:
            mt = (row.start_time - row.maint_time).total_seconds() / 3600
            service_data[wf].append(mt)

    result = {
        "stoppages": [],
        "avg_hours": []
    }

    for wf, types in summary.items():
        for typ, count in types.items():
            result["stoppages"].append({
                "windfarm": wf,
                "type": typ,
                "count": count
            })

    for wf in set(downtime_data.keys()).union(service_data.keys()):
        avg_down = round(sum(downtime_data[wf]) / len(downtime_data[wf]), 2) if downtime_data[wf] else 0
        avg_service = round(sum(service_data[wf]) / len(service_data[wf]), 2) if service_data[wf] else 0
        result["avg_hours"].append({
            "windfarm": wf,
            "avg_downtime_hrs": avg_down,
            "avg_service_hrs": avg_service
        })

    return result

#legend of stoppage
@app.get("/stoppage_legend")
def get_stoppage_legend(
    startdate: str = Query(..., description="Start date in YYYY-MM-DD"),
    enddate: str = Query(..., description="End date in YYYY-MM-DD"),
    db: pyodbc.Connection = Depends(get_db_access)
):
    cursor = db.cursor()

    try:
        start_dt = datetime.strptime(startdate, "%Y-%m-%d")
        end_dt = datetime.strptime(enddate, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    cursor.execute(
        """
        SELECT 
            r.rtnName AS category,
            rr.rsnName AS rsnName
        FROM 
            ((tblEvent AS e
            INNER JOIN tblRationale AS r ON e.rtnID = r.rtnID)
            INNER JOIN tblReason AS rr ON e.rsnID = rr.rsnID)
        WHERE 
            e.dtTS1DownBegin BETWEEN ? AND ?
        """,
        (start_dt, end_dt)
    )

    rows = cursor.fetchall()
    legend_summary = defaultdict(lambda: defaultdict(int))

    for row in rows:
        cat = row.category.strip().lower() if row.category else ""
        rsn = row.rsnName.strip() if row.rsnName else "Unknown"

        if cat == "schedule service":
            typ = "Scheduled Services"
        elif cat in ["fault", "idf fault"]:
            typ = "Faults"
        else:
            typ = "Non Scheduled Services"

        legend_summary[typ][rsn] += 1

    result = []
    for typ, reasons in legend_summary.items():
        for rsn, count in reasons.items():
            result.append({
                "type": typ,
                "rsnName": rsn,
                "count": count
            })
    
    # Sort by count descending
    result.sort(key=lambda x: x["count"], reverse=True)

    return result

# Average RCC response time for faults
@app.get("/get_rcc_response_time")
def get_rcc_response_time(
    startdate: str = Query(..., description="Start date in YYYY-MM-DD"),
    enddate: str = Query(..., description="End date in YYYY-MM-DD"),
    db: pyodbc.Connection = Depends(get_db_access)
):
    cursor = db.cursor()

    try:
        start_dt = datetime.strptime(startdate, "%Y-%m-%d")
        end_dt = datetime.strptime(enddate, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    cursor.execute("""
        SELECT 
            f.facABBR AS windfarm,
            e.dtTS2RCCNotify AS notify_time,
            e.dtTS1DownBegin AS stop_time,
            r.rtnID
        FROM 
            (tblEvent AS e
            INNER JOIN tblFacility AS f ON e.facID = f.facID)
            INNER JOIN tblRationale AS r ON e.rtnID = r.rtnID
        WHERE    
            r.rtnID = 6 AND      
            e.dtTS1DownBegin BETWEEN ? AND ? AND
            e.dtTS2RCCNotify IS NOT NULL AND
            e.dtTS1DownBegin IS NOT NULL
    """, (start_dt, end_dt))

    rows = cursor.fetchall()

    response_time_data = defaultdict(list)

    for row in rows:
        wf = row.windfarm
        if row.stop_time and row.notify_time:
            rt = (row.notify_time - row.stop_time).total_seconds() / 60
            response_time_data[wf].append(rt)

    result = []
    for wf, rtimes in response_time_data.items():
        avg_response = round(sum(rtimes) / len(rtimes), 2) if rtimes else 0
        result.append({
            "windfarm": wf,
            "avg_response_hrs": avg_response
        })

    return {"avg_response_hrs": result}

#Offline Asset heading
@app.get("/offline_headings")
def get_stoppage_legend(
    db: pyodbc.Connection = Depends(get_db_access)
):
    cursor = db.cursor()

    cursor.execute(
        """
        SELECT 
            COUNT(IIf(e.dtTS7EventFinish IS NULL, 1, NULL)) AS total_offline
        FROM 
            (tblEvent AS e
            INNER JOIN tblRationale AS r ON e.rtnID = r.rtnID)
        """
    )

    rows = cursor.fetchall()

    # Assuming the query will return one row with the count result
    if rows:
        total_offline = rows[0][0]
    else:
        total_offline = 0

    # Structure the result into a dictionary or list format
    result = {
        "total_offline": total_offline
    }

    return result

# reading offline wtgs from microsoft access DB
@app.get("/offline_wtgs")
def get_offline_wtgs(db: pyodbc.Connection = Depends(get_db_access)):
    cursor = db.cursor()
    cursor.execute(
        """
        SELECT 
            e.dtTS1DownBegin, 
            f.facABBR, 
            a.astDisplay, 
            r.rtnName, 
            rr.rsnName, 
            n.evntntNote,
            ROUND((IIF(e.dtTS7EventFinish IS NOT NULL, e.dtTS7EventFinish, Now()) - e.dtTS1DownBegin) * 24, 2) AS DowntimeHrs
        FROM 
            ((((tblEvent AS e
            INNER JOIN tblFacility AS f ON e.facID = f.facID)
            INNER JOIN tblAsset AS a ON e.astID = a.astID)
            INNER JOIN tblRationale AS r ON e.rtnID = r.rtnID)
            INNER JOIN tblReason as rr ON e.rsnID = rr.rsnID)
            LEFT JOIN tblEventNotes as n ON e.evntID = n.evntID
        WHERE 
            e.dtTS7EventFinish IS NULL
        ORDER BY  
            f.facABBR ASC,
            a.astDisplay ASC,
            e.dtTS1DownBegin DESC;
        """
        )  # Modify with your actual query
    rows = cursor.fetchall()

    # Extract the column names dynamically from the cursor description
    columns = [column[0] for column in cursor.description]
    
    # Create a list of dictionaries with column names as keys
    data = [dict(zip(columns, row)) for row in rows]

    return {"offlineWtgsDataSet": data}

#Service Details heading
@app.get("/services_details")
def get_services_details(
    startdate: str = Query(..., description="Start date in YYYY-MM-DD"),
    enddate: str = Query(..., description="End date in YYYY-MM-DD"),
    db: pyodbc.Connection = Depends(get_db_access)
):
    cursor = db.cursor()

    try:
        start_dt = datetime.strptime(startdate, "%Y-%m-%d")
        end_dt = datetime.strptime(enddate, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    cursor.execute(
        """
        SELECT 
            COUNT(IIf(r.rtnName NOT IN ('Fault', 'IDF Fault', 'IDF Outage'), 1, NULL)) AS total_services,

            COUNT(IIf(r.rtnName IN ('Schedule Service', 'Scheduled - Adhoc', 'Scheduled Inspections', 'Scheduled Outage'), 1, NULL)) AS scheduled_services,

            COUNT(IIf(
                r.rtnName NOT IN (
                    'Fault', 'IDF Fault', 'IDF Outage', 
                    'Schedule Service', 'Scheduled - Adhoc', 'Scheduled Inspections', 'Scheduled Outage'
                ), 1, NULL)) AS non_scheduled_services,

            ROUND(
                SUM(IIf(r.rtnName NOT IN ('Fault', 'IDF Fault', 'IDF Outage') AND e.dtTS3MaintBegin IS NOT NULL AND e.dtTS7DownFinish IS NOT NULL,
                    DateDiff('s', e.dtTS3MaintBegin, e.dtTS7DownFinish), 0)) / 
                COUNT(IIf(r.rtnName NOT IN ('Fault', 'IDF Fault', 'IDF Outage') AND e.dtTS3MaintBegin IS NOT NULL AND e.dtTS7DownFinish IS NOT NULL, 1, NULL)) / 3600.0
            , 2) AS avg_maint,

            ROUND(
                SUM(IIf(r.rtnName NOT IN ('Fault', 'IDF Fault', 'IDF Outage') AND e.dtTS1DownBegin IS NOT NULL AND e.dtTS7DownFinish IS NOT NULL,
                    DateDiff('s', e.dtTS1DownBegin, e.dtTS7DownFinish), 0)) / 
                COUNT(IIf(r.rtnName NOT IN ('Fault', 'IDF Fault', 'IDF Outage') AND e.dtTS1DownBegin IS NOT NULL AND e.dtTS7DownFinish IS NOT NULL, 1, NULL)) / 3600.0
            , 2) AS avg_down_time
            
            

        FROM 
            tblEvent AS e
            LEFT JOIN tblRationale AS r ON e.rtnID = r.rtnID

        WHERE 
            e.dtTS1DownBegin BETWEEN ? AND ?
        """,
        (start_dt, end_dt)
    )

    row = cursor.fetchone()

    if row:
        total_services = row[0]
        scheduled_services = row[1]
        non_scheduled_services = row[2]
        avg_service_time = row[3]
        avg_down_time = row[4]
    else:
        total_services = 0
        scheduled_services = 0
        non_scheduled_services = 0
        avg_service_time = 0.0
        avg_down_time = 0.0

    return {
        "total_services": total_services,
        "scheduled_services": scheduled_services,
        "non_scheduled_services": non_scheduled_services,
        "avg_service_time": avg_service_time,
        "avg_down_time": avg_down_time
    }

# reading service events from microsoft access DB
@app.get("/get_services")
async def get_services(
    startdate: str = Query(..., description="Start date in format YYYY-MM-DD"),
    enddate: str = Query(..., description="End date in format YYYY-MM-DD"),
    db: pyodbc.Connection = Depends(get_db_access)):
    cursor = db.cursor()
    try:
        start_dt = datetime.strptime(startdate, "%Y-%m-%d")
        end_dt = datetime.strptime(enddate, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        return {"error": "Invalid date format. Use YYYY-MM-DD"}
    except ValueError:
        return {"error": "Invalid date format. Use YYYY-MM-DD"}
    cursor.execute(
        """
        SELECT 
            e.dtTS1DownBegin,  
            e.dtTS7DownFinish,
            ROUND((IIF(e.dtTS7EventFinish IS NOT NULL, e.dtTS7EventFinish, Now()) - e.dtTS1DownBegin) * 24, 2) AS DowntimeHrs,
            f.facABBR, 
            a.astDisplay, 
            r.rtnName, 
            rr.rsnName, 
            n.evntntNote
        FROM 
            ((((tblEvent AS e
            INNER JOIN tblFacility AS f ON e.facID = f.facID)
            INNER JOIN tblAsset AS a ON e.astID = a.astID)
            INNER JOIN tblRationale AS r ON e.rtnID = r.rtnID)
            LEFT JOIN tblReason as rr ON e.rsnID = rr.rsnID)
            LEFT JOIN tblEventNotes as n ON e.evntID = n.evntID
        WHERE 
            e.dtTS1DownBegin BETWEEN ? AND ?
            AND r.rtnName NOT IN ('Fault', 'IDF Outage', 'IDF Fault')
            AND rr.rsnName <> 'Communication loss'
            AND (
                n.evntntNote IS NULL OR n.evntntNote <> 'DELETED'
                )
        ORDER BY 
            f.facABBR ASC,
            e.dtTS1DownBegin DESC,
            a.astDisplay DESC;
        """,
        (start_dt, end_dt)
        )  # Modify with your actual query
    rows = cursor.fetchall()

    # Extract the column names dynamically from the cursor description
    columns = [column[0] for column in cursor.description]
    
    # Create a list of dictionaries with column names as keys
    data = [dict(zip(columns, row)) for row in rows]

    return {"servicesDataSet": data}

#Faults Details heading
@app.get("/faults_details")
def get_faults_details(
    startdate: str = Query(..., description="Start date in YYYY-MM-DD"),
    enddate: str = Query(..., description="End date in YYYY-MM-DD"),
    db: pyodbc.Connection = Depends(get_db_access)
):
    cursor = db.cursor()

    try:
        start_dt = datetime.strptime(startdate, "%Y-%m-%d")
        end_dt = datetime.strptime(enddate, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    cursor.execute(
         """
            SELECT 
                COUNT(*) AS total_faults,

                COUNT(IIf(e.rstbyID = 2, 1, NULL)) AS reset_by_rcc,

                AVG(IIf(e.dtTS7DownFinish IS NOT NULL AND e.dtTS1DownBegin IS NOT NULL,
                    DateDiff('s', e.dtTS1DownBegin, e.dtTS7DownFinish), NULL)) / 3600.0 AS avg_downtime_hrs,

                AVG(IIf(e.dtTS2RCCNotify IS NOT NULL AND e.dtTS1DownBegin IS NOT NULL AND e.dtTS7DownFinish IS NOT NULL,
                    DateDiff('s', e.dtTS1DownBegin, e.dtTS2RCCNotify), NULL)) / 60.0 AS avg_rcc_response_mins

            FROM tblEvent AS e
            INNER JOIN tblRationale AS r ON e.rtnID = r.rtnID
            WHERE 
                r.rtnName = 'Fault' AND
                e.dtTS1DownBegin BETWEEN ? AND ? AND
                e.dtTS7DownFinish IS NOT NULL
            """,
        (start_dt, end_dt)
    )

    row = cursor.fetchone()

    if row:
        result = {
            "total_faults": row[0],
            "reset_by_rcc": row[1],
            "avg_downtime_hrs": round(row[2] or 0, 2),
            "avg_rcc_response_mins": round(row[3] or 0, 2)
        }
    else:
        result = {
            "total_faults": 0,
            "reset_by_rcc": 0,
            "avg_downtime_hrs": 0.0,
            "avg_rcc_response_mins": 0.0
        }

    return result

# reading fault events from microsoft access DB
@app.get("/get_faults")
async def get_faults(
    startdate: str = Query(..., description="Start date in format YYYY-MM-DD"),
    enddate: str = Query(..., description="End date in format YYYY-MM-DD"),
    db: pyodbc.Connection = Depends(get_db_access)):
    cursor = db.cursor()
    try:
        start_dt = datetime.strptime(startdate, "%Y-%m-%d")
        end_dt = datetime.strptime(enddate, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        return {"error": "Invalid date format. Use YYYY-MM-DD"}
    except ValueError:
        return {"error": "Invalid date format. Use YYYY-MM-DD"}
    cursor.execute(
        """
        SELECT 
            f.facABBR, 
            a.astDisplay, 
            r.rtnName, 
            fa.fltCode,
            fa.fltDesc,
            e.dtTS1DownBegin,
            e.dtTS7DownFinish,
            ROUND((IIF(e.dtTS7EventFinish IS NOT NULL, e.dtTS7EventFinish, Now()) - e.dtTS1DownBegin) * 24, 2) AS DowntimeHrs,
            rrt.rsttypName AS ResetType,
            rrb.rstbyName AS ResetBy,
            n.evntntNote
        FROM 
            (((((((tblEvent AS e
            INNER JOIN tblFacility AS f ON e.facID = f.facID)
            INNER JOIN tblAsset AS a ON e.astID = a.astID)
            INNER JOIN tblRationale AS r ON e.rtnID = r.rtnID)
            INNER JOIN tblReason AS rr ON e.rsnID = rr.rsnID)
            INNER JOIN tblFaultCode AS fa ON e.fltID = fa.fltID)
            LEFT JOIN tblRCCResetType AS rrt ON e.rsttypID = rrt.rsttypID)
            LEFT JOIN tblRCCResetBy AS rrb ON e.rstbyID = rrb.rstbyID)
            LEFT JOIN tblEventNotes AS n ON e.evntID = n.evntID
        WHERE 
            e.fltID IS NOT NULL AND
            e.dtTS1DownBegin BETWEEN ? AND ? AND
            e.dtTS7DownFinish IS NOT NULL
        ORDER BY 
            f.facABBR ASC,
            a.astDisplay ASC,
            e.dtTS1DownBegin DESC;
        """,
        (start_dt, end_dt)
        ) 
    rows = cursor.fetchall()

    # Extract the column names dynamically from the cursor description
    columns = [column[0] for column in cursor.description]
    
    # Create a list of dictionaries with column names as keys
    data = [dict(zip(columns, row)) for row in rows]

    return {"faultsDataSet": data}

# @app.get("/overnight_rcc_resets")
# async def get_overnight_rcc_resets(
#     startdate: str = Query(..., description="Start date in format YYYY-MM-DD"),
#     enddate: str = Query(..., description="End date in format YYYY-MM-DD"),
#     db: pyodbc.Connection = Depends(get_db_access),
#     db_prod: pyodbc.Connection = Depends(get_db_prod_stats),
# ):
#     cursor = db.cursor()
#     prod_cursor = db_prod.cursor()

#     try:
#         start_dt = datetime.strptime(startdate, "%Y-%m-%d")
#         end_dt = datetime.strptime(enddate, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
#     except ValueError:
#         return {"error": "Invalid date format. Use YYYY-MM-DD"}

#     cursor.execute(
#         """
#         SELECT 
#             f.facABBR AS WindFarm,
#             a.astDisplay AS WTG,
#             fa.fltCode AS AlarmCode,
#             fa.fltDesc AS AlarmDescription,
#             e.dtTS1DownBegin AS StopTime,               
#             e.dtTS7DownFinish AS StartTime,             
#             ROUND((IIF(e.dtTS1DownBegin IS NOT NULL AND e.dtTS7DownFinish IS NOT NULL, 
#                    (e.dtTS7DownFinish - e.dtTS1DownBegin) * 24, NULL)), 2) AS DowntimeHrs,
#             rrt.rsttypName AS ResetType,
#             n.evntntNote AS Remarks,
#             a.astID
#         FROM 
#             ((((tblEvent AS e
#             INNER JOIN tblFacility AS f ON e.facID = f.facID)
#             INNER JOIN tblAsset AS a ON e.astID = a.astID)
#             INNER JOIN tblFaultCode AS fa ON e.fltID = fa.fltID)
#             LEFT JOIN tblRCCResetType AS rrt ON e.rsttypID = rrt.rsttypID)
#             LEFT JOIN tblEventNotes AS n ON e.evntID = n.evntID
#         WHERE 
#             e.fltID IS NOT NULL AND
#             e.rstbyID = 2 AND
#             e.dtTS1DownBegin BETWEEN ? AND ?
#         ORDER BY 
#             e.dtTS1DownBegin DESC
#         """,
#         (start_dt, end_dt)
#     )

#     rows = cursor.fetchall()
#     columns = [col[0] for col in cursor.description]

#     results = []
#     for row in rows:
#         row_dict = dict(zip(columns, row))

#         stop_time = row_dict["StopTime"]
#         start_time = row_dict["StartTime"]
#         ast_id = row_dict["astID"]
#         stop_date = stop_time.date()

#         # 🕖 Overnight window: 7 PM (previous day) to 7 AM (current day)
#         overnight_start = datetime.combine(stop_date, time(19, 0)) - timedelta(days=1)
#         overnight_end = datetime.combine(stop_date, time(7, 0))

#         # Skip if not in overnight range
#         if not (overnight_start <= stop_time <= overnight_end):
#             continue

#         # ⏱ Calculate saved time
#         saved_time_hrs = max(0, (overnight_end - start_time).total_seconds() / 3600)

#         # ⚡ Fetch daily energy from tblStatsProd (Access)
#         prod_cursor.execute(
#             """
#             SELECT spActEnergyExport FROM tblStatsProd
#             WHERE astID = ? AND spDate = ?
#             """,
#             ast_id, stop_date
#         )
#         prod_row = prod_cursor.fetchone()
#         daily_energy_kwh = prod_row[0] if prod_row else 0
#         daily_energy_mwh = daily_energy_kwh / 1000  # Convert to MWh
#         saved_energy = round((daily_energy_mwh / 24) * saved_time_hrs, 3)

#         # 🧾 Append result row
#         row_dict["SavedTimeHrs"] = round(saved_time_hrs, 2)
#         row_dict["SavedEnergyMWh"] = saved_energy

#         # Remove internal ID
#         del row_dict["astID"]

#         results.append(row_dict)

#     return {"overnightResetsDataSet": results}

@app.get("/overnight_rcc_resets")
async def get_overnight_rcc_resets(
    startdate: str = Query(..., description="Start date in format YYYY-MM-DD"),
    enddate: str = Query(..., description="End date in format YYYY-MM-DD"),
    db: pyodbc.Connection = Depends(get_db_access),
    db_prod: pyodbc.Connection = Depends(get_db_prod_stats),
):
    cursor = db.cursor()
    prod_cursor = db_prod.cursor()

    try:
        start_dt = datetime.strptime(startdate, "%Y-%m-%d")
        end_dt = datetime.strptime(enddate, "%Y-%m-%d")
    except ValueError:
        return {"error": "Invalid date format. Use YYYY-MM-DD"}

    query_start = start_dt - timedelta(days=1)
    query_end = end_dt + timedelta(days=1)

    cursor.execute(
        """
        SELECT 
            f.facABBR AS WindFarm,
            a.astDisplay AS WTG,
            fa.fltCode AS AlarmCode,
            fa.fltDesc AS AlarmDescription,
            e.dtTS1DownBegin AS StopTime,               
            e.dtTS7DownFinish AS StartTime,             
            ROUND((IIF(e.dtTS1DownBegin IS NOT NULL AND e.dtTS7DownFinish IS NOT NULL, 
                   (e.dtTS7DownFinish - e.dtTS1DownBegin) * 24, NULL)), 2) AS DowntimeHrs,
            rrt.rsttypName AS ResetType,
            n.evntntNote AS Remarks,
            a.astID
        FROM 
            ((((tblEvent AS e
            INNER JOIN tblFacility AS f ON e.facID = f.facID)
            INNER JOIN tblAsset AS a ON e.astID = a.astID)
            INNER JOIN tblFaultCode AS fa ON e.fltID = fa.fltID)
            LEFT JOIN tblRCCResetType AS rrt ON e.rsttypID = rrt.rsttypID)
            LEFT JOIN tblEventNotes AS n ON e.evntID = n.evntID
        WHERE 
            e.fltID IS NOT NULL AND
            e.rstbyID = 2 AND
            e.dtTS1DownBegin BETWEEN ? AND ?
        ORDER BY 
            f.facABBR ASC,
            a.astDisplay ASC,
            e.dtTS1DownBegin DESC;
        """,
        (query_start, query_end)
    )

    rows = cursor.fetchall()
    columns = [col[0] for col in cursor.description]

    def daterange(start_date: datetime, end_date: datetime) -> Generator[datetime, None, None]:
        for n in range((end_date - start_date).days + 1):
            yield start_date + timedelta(n)

    results = []
    for row in rows:
        row_dict = dict(zip(columns, row))

        stop_time = row_dict["StopTime"]
        start_time = row_dict["StartTime"]
        ast_id = row_dict["astID"]

        matched = False
        for d in daterange(start_dt, end_dt):
            overnight_start = datetime.combine(d - timedelta(days=1), time(19, 0))  
            overnight_end = datetime.combine(d, time(7, 0)) 

            if overnight_start <= stop_time <= overnight_end:
                matched = True
                break  

        if not matched:
            continue 

        saved_time_hrs = max(0, (overnight_end - start_time).total_seconds() / 3600)

        prod_cursor.execute(
            """
            SELECT spActEnergyExport FROM tblStatsProd
            WHERE astID = ? AND spDate = ?
            """,
            ast_id, d.date()  
        )
        prod_row = prod_cursor.fetchone()
        daily_energy_kwh = prod_row[0] if prod_row else 0
        daily_energy_mwh = daily_energy_kwh / 1000
        saved_energy = round((daily_energy_mwh / 24) * saved_time_hrs, 3)

        row_dict["SavedTimeHrs"] = round(saved_time_hrs, 2)
        row_dict["SavedEnergyMWh"] = saved_energy
        del row_dict["astID"]

        results.append(row_dict)

    return {"overnightResetsDataSet": results}

@app.get("/idf_faults_heading")
def get_idf_faults_heading(
    startdate: str = Query(..., description="Start date in YYYY-MM-DD"),
    enddate: str = Query(..., description="End date in YYYY-MM-DD"),
    db: pyodbc.Connection = Depends(get_db_access)
):
    cursor = db.cursor()

    try:
        start_dt = datetime.strptime(startdate, "%Y-%m-%d")
        end_dt = datetime.strptime(enddate, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    cursor.execute(
        """
        SELECT 
            e.stpID,
            e.rstbyID,
            e.dtTS1DownBegin,
            e.dtTS7DownFinish
        FROM tblEvent AS e
        INNER JOIN tblRationale AS r ON e.rtnID = r.rtnID
        WHERE 
            r.rtnName = 'IDF Fault' AND
            e.dtTS1DownBegin BETWEEN ? AND ?
        """,
        (start_dt, end_dt)
    )

    rows = cursor.fetchall()

    restart_count = 0
    restart_reset_by_rcc = 0
    restart_total_downtime = 0.0
    restart_downtime_count = 0
    restart_total_saving = 0.0

    curtailment_count = 0

    for row in rows:
        stpID = row.stpID
        rstbyID = row.rstbyID
        begin = row.dtTS1DownBegin
        finish = row.dtTS7DownFinish

        if stpID == 434:  # IDF Restart Failure 102
            restart_count += 1

            if begin and finish:
                downtime = (finish - begin).total_seconds() / 3600.0
                restart_total_downtime += downtime
                restart_downtime_count += 1

                if rstbyID == 2:
                    restart_total_saving += max(0, 2 - downtime)

            if rstbyID == 2:
                restart_reset_by_rcc += 1

        elif stpID == 442:  # IDF Curtailment Failure 110
            curtailment_count += 1

    result = {
    "total_idf_faults": len(rows),
    "idf_restart_failures": {
        "count": restart_count,
        "reset_by_rcc": restart_reset_by_rcc,
        "avg_downtime_hrs": round(restart_total_downtime / restart_downtime_count, 2) if restart_downtime_count else 0.0,
        "total_saving_hrs": round(restart_total_saving, 2)
    },
    "idf_curtailment_failures": {
        "count": curtailment_count
    }
}


    return result

@app.get("/get_idf")
async def get_idf(
    startdate: str = Query(..., description="Start date in format YYYY-MM-DD"),
    enddate: str = Query(..., description="End date in format YYYY-MM-DD"),
    db: pyodbc.Connection = Depends(get_db_access)):
    cursor = db.cursor()
    try:
        start_dt = datetime.strptime(startdate, "%Y-%m-%d")
        end_dt = datetime.strptime(enddate, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        return {"error": "Invalid date format. Use YYYY-MM-DD"}
    except ValueError:
        return {"error": "Invalid date format. Use YYYY-MM-DD"}
    cursor.execute(
        """
        SELECT 
            e.dtTS1DownBegin, 
            f.facABBR, 
            a.astDisplay, 
            r.rtnName, 
            s.stpStopDesc,
            ROUND((IIF(e.dtTS7EventFinish IS NOT NULL, e.dtTS7EventFinish, Now()) - e.dtTS1DownBegin) * 24, 2) AS DowntimeHrs,
            rb.rstbyName,
            ROUND((IIF(e.dtTS2RCCNotify IS NOT NULL, e.dtTS2RCCNotify, Now()) - e.dtTS1DownBegin) * 24 * 60, 2) AS ResponseTimeMins,
            ROUND(
                    IIF(
                        r.rtnName = 'IDF Fault' AND 
                        s.stpStopCode = 102 AND 
                        e.rstbyID = 2 AND 
                        ((IIF(e.dtTS7EventFinish IS NOT NULL, e.dtTS7EventFinish, Now()) - e.dtTS1DownBegin) * 24) < 2,
                        2 - ((IIF(e.dtTS7EventFinish IS NOT NULL, e.dtTS7EventFinish, Now()) - e.dtTS1DownBegin) * 24),
                        NULL
                    ),
                    2
                ) AS IDFFaultTimeSaving,
            n.evntntNote
        FROM 
            ((((((tblEvent AS e
            INNER JOIN tblFacility AS f ON e.facID = f.facID)
            INNER JOIN tblAsset AS a ON e.astID = a.astID)
            INNER JOIN tblRationale AS r ON e.rtnID = r.rtnID)
            INNER JOIN tblReason as rr ON e.rsnID = rr.rsnID)
            INNER JOIN tblStopCodes as s ON e.stpID = s.stpID)
            INNER JOIN tblRCCResetBy as rb ON e.rstbyID = rb.rstbyID)
            LEFT JOIN tblEventNotes as n ON e.evntID = n.evntID
        WHERE 
            e.dtTS1DownBegin BETWEEN ? AND ?
            AND (r.rtnName = 'IDF Fault' OR rr.rsnName = 'IDF fault')
            AND n.evntntNote <> 'DELETED'
        ORDER BY 
            e.dtTS1DownBegin DESC,
            f.facABBR ASC,
            a.astDisplay DESC;
        """,
        (start_dt, end_dt)
        )  # Modify with your actual query
    rows = cursor.fetchall()

    # Extract the column names dynamically from the cursor description
    columns = [column[0] for column in cursor.description]
    
    # Create a list of dictionaries with column names as keys
    data = [dict(zip(columns, row)) for row in rows]

    return {"idfDataSet": data}


#----------------------------------------------Analysis Report ------------------------------------------------------------------
# Power Production Analysis
@app.get("/get_production_analysis")
async def get_production_analysis(
    startdate1: str = Query(...),
    enddate1: str = Query(...),
    startdate2: str = Query(...),
    enddate2: str = Query(...),
    db_access: pyodbc.Connection = Depends(get_db_access),
    db_prod_stats: pyodbc.Connection = Depends(get_db_prod_stats)
):
    cursor_access = db_access.cursor()
    cursor_prod_stats = db_prod_stats.cursor()

    def parse_date(d):
        return datetime.strptime(d, "%Y-%m-%d")

    def run_queries(start_dt, end_dt, period_id):
        # Adjust range end
        end_dt = end_dt + timedelta(days=1) - timedelta(seconds=1)

        # Query 1: Downtime
        query_1 = """
        SELECT 
            DatePart('ww', CDate(tblEvent.dtTS1DownBegin)) AS weeknum, 
            tblFacility.facABBR, 
            COUNT(tblEvent.evntID) AS totalEntries, 
            SUM(ROUND(
                (IIF(tblEvent.dtTS7EventFinish IS NOT NULL, tblEvent.dtTS7EventFinish, Now()) - tblEvent.dtTS1DownBegin) * 24, 2
            )) AS DowntimeHrs 
        FROM tblFacility 
        INNER JOIN tblEvent ON tblFacility.facID = tblEvent.facID 
        WHERE tblEvent.dtTS1DownBegin BETWEEN ? AND ? 
        GROUP BY DatePart('ww', CDate(tblEvent.dtTS1DownBegin)), tblFacility.facABBR
        """

        # Query 2: Power Stats
        query_2 = """
        SELECT 
            DatePart('ww', CDate(tblStatsProd.spDate)) AS weeknum, 
            tblFacility.facABBR, 
            SUM(tblStatsProd.spActEnergyExport) AS TotalExpPower, 
            ROUND(AVG(tblStatsProd.spAvgWS), 2) AS avgWS
        FROM tblFacility 
        INNER JOIN tblStatsProd ON tblFacility.facID = tblStatsProd.facID
        WHERE tblStatsProd.spDate BETWEEN ? AND ? 
        GROUP BY DatePart('ww', CDate(tblStatsProd.spDate)), tblFacility.facABBR
        """

        cursor_access.execute(query_1, (start_dt, end_dt))
        rows_1 = cursor_access.fetchall()
        cursor_prod_stats.execute(query_2, (start_dt, end_dt))
        rows_2 = cursor_prod_stats.fetchall()

        columns_1 = [column[0] for column in cursor_access.description]
        columns_2 = [column[0] for column in cursor_prod_stats.description]

        data_1 = [dict(zip(columns_1, row)) for row in rows_1]
        data_2 = [dict(zip(columns_2, row)) for row in rows_2]

        combined_data = []

        for row_1 in data_1:
            match = next(
                (row_2 for row_2 in data_2 if row_2['weeknum'] == row_1['weeknum'] and row_2['facABBR'] == row_1['facABBR']),
                None
            )
            merged = {**row_1, **(match or {}), "period": period_id}
            combined_data.append(merged)

        return combined_data

    # Run for both time periods
    dt1_start = parse_date(startdate1)
    dt1_end = parse_date(enddate1)
    dt2_start = parse_date(startdate2)
    dt2_end = parse_date(enddate2)

    result1 = run_queries(dt1_start, dt1_end, 'Period 1')
    result2 = run_queries(dt2_start, dt2_end, 'Period 2')

    return {"productionAnalysisDataSet": result1 + result2}

#Scheduled Service Analysis
@app.get("/get_schedule_service_analysis")
def get_schedule_service_analysis(
    startdate1: str = Query(..., description="Start date for Period 1 (YYYY-MM-DD)"),
    enddate1: str = Query(..., description="End date for Period 1 (YYYY-MM-DD)"),
    startdate2: str = Query(..., description="Start date for Period 2 (YYYY-MM-DD)"),
    enddate2: str = Query(..., description="End date for Period 2 (YYYY-MM-DD)"),
    db: pyodbc.Connection = Depends(get_db_access)
):
    def run_query(start, end, period):
        try:
            start_dt = datetime.strptime(start, "%Y-%m-%d")
            end_dt = datetime.strptime(end, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format")

        cursor = db.cursor()
        cursor.execute(
            """
            SELECT  
                f.facABBR,
                a.astDisplay,  
                r.rtnName,
                rr.rsnName,
                ROUND(SUM(
                    IIF(e.dtTS7DownFinish IS NOT NULL, 
                        (e.dtTS7DownFinish - e.dtTS1DownBegin) * 24, 
                        (Now() - e.dtTS1DownBegin) * 24
                    )
                ), 2) AS total_downtime_hrs,
                e.dtTS1DownBegin
            FROM 
                ((((tblEvent AS e
                INNER JOIN tblFacility AS f ON e.facID = f.facID)
                INNER JOIN tblAsset AS a ON e.astID = a.astID)
                INNER JOIN tblRationale AS r ON e.rtnID = r.rtnID)
                LEFT JOIN tblReason as rr ON e.rsnID = rr.rsnID)
                LEFT JOIN tblEventNotes as n ON e.evntID = n.evntID
            WHERE 
                e.dtTS1DownBegin BETWEEN ? AND ?
                AND r.rtnName = 'Schedule Service'
                AND rr.rsnName <> 'Communication loss'
                AND n.evntntNote <> 'DELETED'
            GROUP BY
                f.facABBR, a.astDisplay, r.rtnName, rr.rsnName, e.dtTS1DownBegin
            ORDER BY 
                f.facABBR ASC
            """,
            (start_dt, end_dt)
        )

        temp_data = defaultdict(lambda: {
            "count": 0,
            "total_downtime_hrs": 0.0,
            "week_number": None
        })

        for row in cursor.fetchall():
            wind_farm = row[0]
            wtg = row[1]
            rtn_name = row[2]
            rsn_name = row[3]
            downtime = float(row[4]) if row[4] is not None else 0.0
            down_begin = row[5]

            try:
                if isinstance(down_begin, str):
                    down_begin = datetime.fromisoformat(down_begin)
                week_number = down_begin.isocalendar().week
            except Exception:
                week_number = None

            key = (period, wind_farm, wtg, rtn_name, rsn_name)
            temp_data[key]["count"] += 1
            temp_data[key]["total_downtime_hrs"] += downtime
            temp_data[key]["week_number"] = week_number

        result = []
        for (period, wind_farm, wtg, rtn_name, rsn_name), values in temp_data.items():
            count = values["count"]
            total_downtime = values["total_downtime_hrs"]
            avg_downtime = round(total_downtime / count, 2) if count > 0 else 0.0

            result.append({
                "period": period,
                "wind_farm": wind_farm,
                "rtn_name": rtn_name,
                "rsn_name": rsn_name,
                "wtg": wtg,
                "count": count,
                "avg_downtime_hrs": avg_downtime,
                "total_downtime_hrs": round(total_downtime, 2),
                "week_number": values["week_number"]
            })

        return result

    results = run_query(startdate1, enddate1, 'period 1') + run_query(startdate2, enddate2, 'period 2')
    results.sort(key=lambda x: (x["wind_farm"], x["wtg"], x["rtn_name"], x["rsn_name"]))  # Sort by wind farm and more


    return {
        "scheduledserviceAnalysisDataSet": results
    }

#All Other Services
@app.get("/get_service_analysis")
def get_service_analysis(
    startdate1: str = Query(..., description="Start date for Period 1 (YYYY-MM-DD)"),
    enddate1: str = Query(..., description="End date for Period 1 (YYYY-MM-DD)"),
    startdate2: str = Query(..., description="Start date for Period 2 (YYYY-MM-DD)"),
    enddate2: str = Query(..., description="End date for Period 2 (YYYY-MM-DD)"),
    db: pyodbc.Connection = Depends(get_db_access)
):
    def run_query(start, end, period):
        try:
            start_dt = datetime.strptime(start, "%Y-%m-%d")
            end_dt = datetime.strptime(end, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format")

        cursor = db.cursor()
        cursor.execute(
            """
            SELECT  
                f.facABBR,  
                r.rtnName, 
                COUNT(*) AS frequency,
                ROUND(SUM(
                    IIF(e.dtTS7DownFinish IS NOT NULL, 
                        (e.dtTS7DownFinish - e.dtTS1DownBegin) * 24, 
                        (Now() - e.dtTS1DownBegin) * 24
                    )
                ), 2) AS total_downtime_hrs,
                e.dtTS1DownBegin
            FROM 
                ((((tblEvent AS e
                INNER JOIN tblFacility AS f ON e.facID = f.facID)
                INNER JOIN tblAsset AS a ON e.astID = a.astID)
                INNER JOIN tblRationale AS r ON e.rtnID = r.rtnID)
                LEFT JOIN tblReason as rr ON e.rsnID = rr.rsnID)
                LEFT JOIN tblEventNotes as n ON e.evntID = n.evntID
            WHERE 
                e.dtTS1DownBegin BETWEEN ? AND ?
                AND r.rtnName NOT IN ('Fault', 'IDF Outage', 'Other', 'IDF Fault', 'Schedule Service', 'Schedule Outage')
                AND rr.rsnName <> 'Communication loss'
                AND n.evntntNote <> 'DELETED'
            GROUP BY
                f.facABBR, r.rtnName, e.dtTS1DownBegin
            ORDER BY 
                f.facABBR ASC,
                COUNT(*) DESC
            """,
            (start_dt, end_dt)
        )

        rows = cursor.fetchall()
        result = []
        for row in rows:
            down_begin = row[4]
            try:
                if isinstance(down_begin, str):
                    down_begin = datetime.fromisoformat(down_begin)
                week_number = down_begin.isocalendar().week
            except Exception:
                week_number = None

            # Add each result to the final list
            result.append({
                "week_number": week_number,
                "wind_farm": row[0],
                "rationale": row[1],
                "count": row[2],
                "total_downtime_hrs": row[3],
                "period": period
            })
        return result

    # Run queries for both date ranges and merge the results
    results = run_query(startdate1, enddate1, 'period 1') + run_query(startdate2, enddate2, 'period 2')

    # Now, combine the results by week_number, wind_farm, and rationale
    combined_results = {}
    for entry in results:
        key = (entry["week_number"], entry["wind_farm"], entry["rationale"])

        if key not in combined_results:
            combined_results[key] = {
                "week_number": entry["week_number"],
                "wind_farm": entry["wind_farm"],
                "rationale": entry["rationale"],
                "count": 0,
                "total_downtime_hrs": 0.0,
                "period": entry["period"],
            }

        # Combine the count and downtime values for matching entries
        combined_results[key]["count"] += entry["count"]
        combined_results[key]["total_downtime_hrs"] += entry["total_downtime_hrs"]

    # Convert the combined results back into a list
    final_results = list(combined_results.values())
    final_results.sort(key=lambda x: (x["wind_farm"], x["rationale"]))  # Sort by wind_farm, then rationale


    return {
        "serviceAnalysisDataSet": final_results
    }

# Top 10 Faults Analysis
@app.get("/top_ten_faults")
def get_top_ten_faults(
    startdate1: str = Query(..., description="Start date for Period 1 YYYY-MM-DD"),
    enddate1: str = Query(..., description="End date for Period 1 YYYY-MM-DD"),
    startdate2: str = Query(..., description="Start date for Period 2 YYYY-MM-DD"),
    enddate2: str = Query(..., description="End date for Period 2 YYYY-MM-DD"),
    sort_by: str = Query("downtime", description="Sort by 'downtime' or 'frequency'"),
    db: pyodbc.Connection = Depends(get_db_access)
):
    def run_query(start, end, sort_by):
        try:
            start_dt = datetime.strptime(start, "%Y-%m-%d")
            end_dt = datetime.strptime(end, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format")

        if sort_by == "downtime":
            order_by = """
            SUM(
                IIF(e.dtTS7DownFinish IS NOT NULL, 
                    (e.dtTS7DownFinish - e.dtTS1DownBegin) * 24, 
                    (Now() - e.dtTS1DownBegin) * 24
                )
            ) DESC
            """
        elif sort_by == "frequency":
            order_by = "COUNT(*) DESC"
        else:
            raise HTTPException(status_code=400, detail="Invalid sort_by value. Use 'downtime' or 'frequency'.")

        cursor = db.cursor()
        cursor.execute(
            f"""
            SELECT 
                TOP 10
                fa.fltCode, 
                fa.fltDesc, 
                COUNT(*) AS frequency,
                SUM(
                    IIF(e.dtTS7DownFinish IS NOT NULL, 
                        (e.dtTS7DownFinish - e.dtTS1DownBegin) * 24, 
                        (Now() - e.dtTS1DownBegin) * 24
                    )
                ) AS total_downtime_hrs,
                MIN(e.dtTS1DownBegin) AS first_occurrence
            FROM 
                (tblEvent AS e
                INNER JOIN tblRationale AS r ON e.rtnID = r.rtnID)
                INNER JOIN tblFaultCode AS fa ON e.fltID = fa.fltID
            WHERE 
                r.rtnName = 'Fault'
                AND e.dtTS1DownBegin BETWEEN ? AND ?
            GROUP BY 
                fa.fltCode, fa.fltDesc
            ORDER BY 
                {order_by},
                MIN(e.dtTS1DownBegin) ASC
            """,
            (start_dt, end_dt)
        )

        rows = cursor.fetchall()

        result = []
        for row in rows:
            result.append({
                "fault_code": row[0],
                "description": row[1],
                "count": row[2],
                "total_downtime_hrs": round(row[3], 2) if row[3] is not None else 0.0
            })

        return result

    period1_results = run_query(startdate1, enddate1, sort_by)
    period2_results = run_query(startdate2, enddate2, sort_by)

    return {
        "period1": period1_results,
        "period2": period2_results
    }


#Wind Farm Reports and Event Log Data --------------------
@app.get("/prod_stats_by_site")
def get_prod_stats_by_site(
    facid: int = Query(..., description="Facility ID (e.g. 8 for SYHWF)"),
    startdate: str = Query(..., description="Start date in YYYY-MM-DD"),
    enddate: str = Query(..., description="End date in YYYY-MM-DD"),
    db: pyodbc.Connection = Depends(get_db_prod_stats)
):
    try:
        start_dt = datetime.strptime(startdate, "%Y-%m-%d")
        end_dt = datetime.strptime(enddate, "%Y-%m-%d")

        cursor = db.cursor()
        date_column = 'spDate'

        sql = f"""
            SELECT AVG(spAvgWS) AS avg_wind_speed,
                   SUM(spActEnergyExport)/1000.0 AS production_mwh
            FROM tblStatsProd
            WHERE facID = ?
              AND [{date_column}] BETWEEN ? AND ?
        """

        cursor.execute(sql, (facid, start_dt, end_dt))
        row = cursor.fetchone()

        if not row or row[0] is None:
            return {
                "message": f"No data for facID = {facid} between {startdate} and {enddate}"
            }

        return {
            "facID": facid,
            "AVG Wind Speed": round(row[0], 2),
            "Total Production (mWH)": round(row[1], 2)
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# reading offline wtgs for any wind farm
@app.get("/offline_wtgs_for_wf")
def get_offline_wtgs_for_wf(
    windfarm: str = Query(default=None, description="Filter by wind farm abbreviation"),
    db: pyodbc.Connection = Depends(get_db_access)
 ):
    cursor = db.cursor()

    # Build dynamic WHERE clause based on windfarm filter
    query = """
        SELECT 
            e.dtTS1DownBegin, 
            f.facABBR, 
            a.astDisplay, 
            r.rtnName, 
            rr.rsnName, 
            n.evntntNote,
            ROUND((IIF(e.dtTS7EventFinish IS NOT NULL, e.dtTS7EventFinish, Now()) - e.dtTS1DownBegin) * 24, 2) AS DowntimeHrs
        FROM 
            ((((tblEvent AS e
            INNER JOIN tblFacility AS f ON e.facID = f.facID)
            INNER JOIN tblAsset AS a ON e.astID = a.astID)
            INNER JOIN tblRationale AS r ON e.rtnID = r.rtnID)
            INNER JOIN tblReason as rr ON e.rsnID = rr.rsnID)
            LEFT JOIN tblEventNotes as n ON e.evntID = n.evntID
        WHERE 
            e.dtTS7EventFinish IS NULL
            AND n.evntntNote <> 'DELETED'
    """

    params = []

    if windfarm:
        query += " AND f.facABBR = ?"
        params.append(windfarm)

    query += """
        ORDER BY 
            e.dtTS1DownBegin DESC,
            f.facABBR ASC,
            a.astDisplay DESC
    """

    cursor.execute(query, params)
    rows = cursor.fetchall()

    # Extract column names
    columns = [column[0] for column in cursor.description]
    data = [dict(zip(columns, row)) for row in rows]

    return {"offlineWtgsWFDataSet": data}

# get stoppages for any wind farm
@app.get("/get_stoppages_for_wf")
async def get_services(
    startdate: str = Query(..., description="Start date in format YYYY-MM-DD"),
    enddate: str = Query(..., description="End date in format YYYY-MM-DD"),
    windfarm: str = Query(None, description="Wind farm abbreviation (optional)"),
    db: pyodbc.Connection = Depends(get_db_access)
):
    cursor = db.cursor()
    try:
        start_dt = datetime.strptime(startdate, "%Y-%m-%d")
        end_dt = datetime.strptime(enddate, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
    except ValueError:
        return {"error": "Invalid date format. Use YYYY-MM-DD"}
    
    # Build the query dynamically based on the presence of the windfarm filter
    query = """
    SELECT 
        f.facABBR, 
        a.astDisplay, 
        e.evntWindSpeed,
        r.rtnName, 
        rr.rsnName, 
        s.stpStopCode,
        s.stpStopDesc,
        fa.fltCode,
        fa.fltDesc,
        ROUND((IIF(e.dtTS7DownFinish IS NOT NULL, e.dtTS7DownFinish, Now()) - e.dtTS1DownBegin) * 24, 2) AS DowntimeHrs,
        e.dtTS1DownBegin, 
        e.dtTS2RCCNotify,
        e.dtTS7DownFinish,
        e.dtTS3MaintBegin,
        rt.rsttypName,
        rb.rstbyName,
        n.evntntNote
    FROM 
        ((((((((tblEvent AS e
        INNER JOIN tblFacility AS f ON e.facID = f.facID)
        INNER JOIN tblAsset AS a ON e.astID = a.astID)
        LEFT JOIN tblRationale AS r ON e.rtnID = r.rtnID)
        INNER JOIN tblReason as rr ON e.rsnID = rr.rsnID)
        INNER JOIN tblStopCodes as s ON e.stpID = s.stpID)
        LEFT JOIN tblFaultCode as fa ON e.fltID = fa.fltID)
        LEFT JOIN tblRCCResetType as rt ON e.rsttypID = rt.rsttypID)
        LEFT JOIN tblRCCResetBy as rb ON e.rstbyID = rb.rstbyID)
        LEFT JOIN tblEventNotes as n ON e.evntID = n.evntID
    WHERE
        e.dtTS1DownBegin BETWEEN ? AND ?
        AND (n.evntntNote <> 'DELETED' OR n.evntntNote IS NULL OR n.evntntNote = '')
    """
    
    # If a windfarm is provided, add it as a filter in the query
    if windfarm:
        query += " AND f.facABBR = ?"
    
    # Execute the query, passing the appropriate parameters
    if windfarm:
        cursor.execute(query, (start_dt, end_dt, windfarm))
    else:
        cursor.execute(query, (start_dt, end_dt))
    
    rows = cursor.fetchall()

    # Extract the column names dynamically from the cursor description
    columns = [column[0] for column in cursor.description]
    
    # Create a list of dictionaries with column names as keys
    data = [dict(zip(columns, row)) for row in rows]

    return {"stoppagesDataSet": data}

#get idf for analysis
@app.get("/get_idf_analysis")
async def get_idf_analysis(
    startdate1: str = Query(..., description="Start date for Period 1 (YYYY-MM-DD)"),
    enddate1: str = Query(..., description="End date for Period 1 (YYYY-MM-DD)"),
    startdate2: str = Query(..., description="Start date for Period 2 (YYYY-MM-DD)"),
    enddate2: str = Query(..., description="End date for Period 2 (YYYY-MM-DD)"),
    db: pyodbc.Connection = Depends(get_db_access)
):
    def run_query(start, end, period):
        try:
            start_dt = datetime.strptime(start, "%Y-%m-%d")
            end_dt = datetime.strptime(end, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format")

        cursor = db.cursor()
        cursor.execute(
            """
            SELECT 
                e.dtTS1DownBegin, 
                f.facABBR, 
                a.astDisplay, 
                r.rtnName, 
                s.stpStopDesc,
                ROUND((IIF(e.dtTS7EventFinish IS NOT NULL, e.dtTS7EventFinish, Now()) - e.dtTS1DownBegin) * 24, 2) AS DowntimeHrs,
                rb.rstbyName,
                ROUND((IIF(e.dtTS2RCCNotify IS NOT NULL, e.dtTS2RCCNotify, Now()) - e.dtTS1DownBegin) * 24 * 60, 2) AS ResponseTimeMins,
                ROUND(
                    IIF(
                        r.rtnName = 'IDF Fault' AND 
                        s.stpStopCode = 102 AND 
                        e.rstbyID = 2 AND 
                        ((IIF(e.dtTS7EventFinish IS NOT NULL, e.dtTS7EventFinish, Now()) - e.dtTS1DownBegin) * 24) < 2,
                        2 - ((IIF(e.dtTS7EventFinish IS NOT NULL, e.dtTS7EventFinish, Now()) - e.dtTS1DownBegin) * 24),
                        NULL
                    ),
                    2
                ) AS IDFFaultTimeSaving,
                n.evntntNote
            FROM 
                ((((((tblEvent AS e
                INNER JOIN tblFacility AS f ON e.facID = f.facID)
                INNER JOIN tblAsset AS a ON e.astID = a.astID)
                INNER JOIN tblRationale AS r ON e.rtnID = r.rtnID)
                INNER JOIN tblReason as rr ON e.rsnID = rr.rsnID)
                INNER JOIN tblStopCodes as s ON e.stpID = s.stpID)
                INNER JOIN tblRCCResetBy as rb ON e.rstbyID = rb.rstbyID)
                LEFT JOIN tblEventNotes as n ON e.evntID = n.evntID
            WHERE 
                e.dtTS1DownBegin BETWEEN ? AND ?
                AND (r.rtnName = 'IDF Fault' OR rr.rsnName = 'IDF fault')
                AND n.evntntNote <> 'DELETED'
            ORDER BY 
                a.astDisplay DESC;
            """,
            (start_dt, end_dt)
        )

        rows = cursor.fetchall()
        columns = [column[0] for column in cursor.description]
        data = [dict(zip(columns, row)) for row in rows]

        # Add period to each entry
        for entry in data:
            entry["period"] = period

        return data

    # Run queries for both date ranges and merge the results
    period1_data = run_query(startdate1, enddate1, 'period 1')
    period2_data = run_query(startdate2, enddate2, 'period 2')

    # Combine the data from both periods
    combined_data = period1_data + period2_data

    return {"idfAnalysisDataSet": combined_data}



# reading from excel (raw data 2025)
# @app.get("/read-excel/", response_model=List[models.ExcelRow])
# async def read_excel():
#     excel_file_path = r"C:\Users\gwarcc\goldwindaustralia\Service SharePoint - Service Technical Library\22 RCC\RCC\18. RCC Reporting\01 Yearly Raw Data\2025\RCC Benefit Raw Data 2025.xlsm"

#     wb = load_workbook(excel_file_path)
#     sheet = wb.active

#     headers = [
#         "Date", "Wind Farm", "WTG", "WTG Type", "WTG Type 2", "Wind Speed", "Category", 
#         "Reason", "Alarm Code", "Alarm Description", "Downtime", "Stop Time", "Maint Time", 
#         "Start Time", "Remarks", "RCC Notified Time", "Before or After RCC Control", 
#         "Weekend Day/Hour", "Day/Night", "Reset Level", "RCC Notified time (min)", 
#         "Reset By", "Response Time", "Before reset by Site/ After Reset by RCC", 
#         "IDF Fault Time Saving"
#     ]

#     # Read rows from the Excel sheet and store them in a list of dictionaries
#     rows = []
#     for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
#         row_data = {headers[i]: row[i].value for i in range(len(headers))}
#         rows.append(row_data)

#     return rows
    
# @app.get("/top_fault_codes_detailed")
# def get_top_fault_codes_detailed(
#     startdate1: str = Query(..., description="Start date for Period 1 YYYY-MM-DD"),
#     enddate1: str = Query(..., description="End date for Period 1 YYYY-MM-DD"),
#     startdate2: str = Query(..., description="Start date for Period 2 YYYY-MM-DD"),
#     enddate2: str = Query(..., description="End date for Period 2 YYYY-MM-DD"),
#     db: pyodbc.Connection = Depends(get_db_access)
# ):
#     def run_query(start, end):
#         try:
#             start_dt = datetime.strptime(start, "%Y-%m-%d")
#             end_dt = datetime.strptime(end, "%Y-%m-%d") + timedelta(days=1) - timedelta(seconds=1)
#         except ValueError:
#             raise HTTPException(status_code=400, detail="Invalid date format")

#         cursor = db.cursor()
#         cursor.execute(
#             """
#             SELECT 
#                 TOP 10
#                 fa.fltCode, 
#                 fa.fltDesc, 
#                 COUNT(*) AS frequency,
#                 SUM(
#                     IIF(e.dtTS7DownFinish IS NOT NULL, 
#                         (e.dtTS7DownFinish - e.dtTS1DownBegin) * 24, 
#                         (Now() - e.dtTS1DownBegin) * 24
#                     )
#                 ) AS total_downtime_hrs,
#                 MIN(e.dtTS1DownBegin) AS first_occurrence
#             FROM 
#                 (tblEvent AS e
#                 INNER JOIN tblRationale AS r ON e.rtnID = r.rtnID)
#                 INNER JOIN tblFaultCode AS fa ON e.fltID = fa.fltID
#             WHERE 
#                 r.rtnName = 'Fault'
#                 AND e.dtTS1DownBegin BETWEEN ? AND ?
#             GROUP BY 
#                 fa.fltCode, fa.fltDesc
#             ORDER BY 
#                 COUNT(*) DESC, 
#                 MIN(e.dtTS1DownBegin) ASC
#             """,
#             (start_dt, end_dt)
#         )

#         rows = cursor.fetchall()

#         result = []
#         for row in rows:
#             down_begin = row[4]
#             try:
#                 if isinstance(down_begin, str):
#                     down_begin = datetime.fromisoformat(down_begin)
#                 week_number = down_begin.isocalendar().week
#             except Exception:
#                 week_number = None

#             result.append({
#                 "week_number": week_number,
#                 "fault_code": row[0],
#                 "description": row[1],
#                 "count": row[2],
#                 "total_downtime_hrs": round(row[3], 2) if row[3] is not None else 0.0
#             })

#         return result

#     period1_results = run_query(startdate1, enddate1)
#     period2_results = run_query(startdate2, enddate2)

#     combined = period1_results + period2_results

#     return {"faultCodesDataSet": combined}


  



